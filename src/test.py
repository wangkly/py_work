#个险
from openpyxl import Workbook
from openpyxl import load_workbook
from gongsi import gongsi_filter
from options_filter import outer_filter
from functools import reduce
from rowsum import rowSum
from qimo import qimo

# 实例化
# wb = Workbook()
# 激活 worksheet
# ws = wb.active

workbook = load_workbook('/Users/wangkly/py_work/xuTest/src/辅助余额表.xlsx')
worksheet  = workbook['10月']

#第0行，标题行
row0 = [item.value for item in list(worksheet.rows)[0]]
gonsi_index = row0.index('公司代码')
# print('公司代码index',gonsi_index)

#公司代码 筛选 1900
gongsiFilterd=list(filter(gongsi_filter(gonsi_index,'1900'),list(worksheet.rows)))

#筛科目
kemuIndex = row0.index('科目')
# names=['保费收入－首年首期','保费收入－首年续期','保费收入-保全','保费收入-团体短期意外险保费收入']
kemuOptins =['6031010001','6031010002','6031030000','6031050000']
kemuFilterd = list(filter(outer_filter(kemuOptins,kemuIndex),gongsiFilterd))
# print(len(kemuFilterd))

# print('利润中心 index',row0.index('利润中心'))
#利润中心描述筛选 个险
lirunIndex = row0.index('利润中心')
lirunOptins=['PC10'] #个险
lirunFilterd = list(filter(outer_filter(lirunOptins,lirunIndex),kemuFilterd))

xianzhongIndex = row0.index('险种大类')
#险种大类描述 筛选 短期费补医疗健康 短期普疾健康
#names=['短期费补医疗健康','短期普疾健康','短期定返医疗健康','短期重疾健康']
xianzhongOptions=['18','17','19','16']
xianzhongFilterd =  list(filter(outer_filter(xianzhongOptions,xianzhongIndex),lirunFilterd))

#先map取每一行期间借方，期间贷方之和list
rowCollect = map(rowSum,xianzhongFilterd)

#短期健康险
total = sum(list(rowCollect))
print('短期健康险',total)

#本年保费收入 G6
G6 =  sum(list(map(qimo,xianzhongFilterd)))
print('G6',G6)

#短期健康险 期缴
#缴费方式index
jiaofeiIndex = row0.index('缴费方式')
qijiaoOptions=['2']
qijiao  = list(filter(outer_filter(qijiaoOptions,jiaofeiIndex),xianzhongFilterd))
qijiaoTotal = sum(list(map(rowSum,qijiao)))
print('短期健康-期缴',qijiaoTotal)

#本年保费收入 首年 期缴
H6 = sum(list(map(qimo,qijiao)))
print('H6',H6)


#短期健康险-续期 科目：保费收入续期
kemuOptions2=['6031020000']
xuqiFilterd = list(filter(outer_filter(kemuOptions2,kemuIndex),gongsiFilterd))
lirunFilterd2 = list(filter(outer_filter(lirunOptins,lirunIndex),xuqiFilterd))
xianzhongFilterd2 =  list(filter(outer_filter(xianzhongOptions,xianzhongIndex),lirunFilterd2))
xuqiTotal = sum(list(map(rowSum,xianzhongFilterd2)))
print('短期健康-续期',xuqiTotal)

#本年保费收入 续期
I6=sum(list(map(qimo,xianzhongFilterd2)))
print('I6',I6)



#意外伤害险
yiwaiOptins=['9']
yiwaiFilterd = list(filter(outer_filter(yiwaiOptins,xianzhongIndex),lirunFilterd))
#意外险
yiwaiTotal = sum(list(map(rowSum,yiwaiFilterd)))
print('yiwaiTotal',yiwaiTotal)
#本年保费收入 G7
G7=sum(list(map(qimo,yiwaiFilterd)))
print('G7',G7)

#意外险 期缴
yiwaiQijiao = list(filter(outer_filter(qijiaoOptions,jiaofeiIndex),yiwaiFilterd))
D7 = sum(list(map(rowSum,yiwaiQijiao)))
print('D7',D7)
#本年保费收入 H7
H7= sum(list(map(qimo,yiwaiQijiao)))
print('H7',H7)

#意外险 续期
yiwaiXQFilterd =  list(filter(outer_filter(yiwaiOptins,xianzhongIndex),lirunFilterd2))
#意外险 续期 本月
yiwaiXQ =  sum(list(map(rowSum,yiwaiXQFilterd)))
print('E7',yiwaiXQ)

#意外 续期 本年
yiwaiXQY = sum(list(map(qimo,yiwaiXQFilterd)))
print('I7',yiwaiXQY)

#一般寿险
#普通定期寿险：1，普通两全寿险：2，普通年金寿险：23，普通养老年金寿险：25，普通终生寿险：3，长期定返医疗健康：13，长期普疾健康：11,长期重疾健康：10
putongOptins=['1','2','23','25','3','13','11','10']
yibanFilterd = list(filter(outer_filter(putongOptins,xianzhongIndex),lirunFilterd))

yibanTotal = sum(list(map(rowSum,yibanFilterd)))
print('一般寿险',yibanTotal)

#一般寿险 本年 G8
yibanY =  sum(list(map(qimo,yibanFilterd)))
print('G8',yibanY)

#一般寿险 期缴
yibanQJ = list(filter(outer_filter(qijiaoOptions,jiaofeiIndex),yibanFilterd))
D8 =  sum(list(map(rowSum,yibanQJ)))
print('D8',D8)
#一般寿险 期缴 本年
H8= sum(list(map(qimo,yibanQJ)))
print('H8',H8)

#一般寿险 续期
yibanXQFilterd = list(filter(outer_filter(putongOptins,xianzhongIndex),lirunFilterd2))
yibanXQ = sum(list(map(rowSum,yibanXQFilterd)))
print('E8',yibanXQ)
#一般寿险 续期 本年
yibanXQY = sum(list(map(qimo,yibanXQFilterd)))
print('I8',yibanXQY)


# #保存到tempate.xlsx
# template = load_workbook('/Users/wangkly/py_work/xuTest/src/template.xlsx')

# nanjing =  template['南京']

# nanjing['C6'] = total
# nanjing['D6'] = qijiaoTotal

# template.save('test.xlsx')


