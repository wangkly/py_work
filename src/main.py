import string
from decimal import *
getcontext().prec = 28
from compute import compute
from compute_jifu import computeJF
from openpyxl import Workbook
from openpyxl import load_workbook
from write_file import writeFile

# wb = load_workbook('./src/辅助余额表.xlsx')
# template = load_workbook('./src/template.xlsx')

def entry(workbook,dest,name,gongsi):
    codeDict={'1900':'南京','1902':'盐城','1903':'南通','1904':'无锡','1905':'徐州','1906':'常州','1999':'苏州','-1':'江苏'}
    worksheet = workbook[str(name)]
    template = load_workbook(str(dest))
    #保费收入（正的就是负的，负的就是正的）
    d1 = compute(worksheet,['PC10'],gongsi)
    writeFile(template,d1,0)
    d2 = compute(worksheet,['PC20'],gongsi)
    writeFile(template,d2,1)
    d3 = compute(worksheet,['PC30','PC40'],gongsi)
    writeFile(template,d3,2)

    #计算理赔给付 ，各类给付，退保金额
    d4 = computeJF(worksheet,['PC10'],gongsi)
    writeFile(template,d4,0)
    d5 = computeJF(worksheet,['PC20'],gongsi)
    writeFile(template,d5,1)
    d6 = computeJF(worksheet,['PC30','PC40'],gongsi)
    writeFile(template,d6,2)


    #计算合计
    targetSheet = template['南京'] #ws
    # targetSheet = template[str(gongsi)] #ws
    columns = list(string.ascii_uppercase[2:16])
    columns.append('T')
    rows = [12,19,26]
    for column in columns:
        aTotal = Decimal('0')
        for row in rows:
            akey = str(column)+str(row)
            aTotal +=  Decimal(str( targetSheet[akey].value ))  #C6，C7,C8,C9,C10,C11 
        targetSheet[str(column)+'29'] = aTotal

    # 短险赔付率
    getcontext().prec = 4
    Q6 = (Decimal(str(targetSheet['L6'].value))  / Decimal(str(targetSheet['J6'].value))) if targetSheet['J6'].value != 0 else 0
    targetSheet['Q6'] = str(Q6 * Decimal(100))+'%'
    
    Q7 = (Decimal(str(targetSheet['L7'].value)) / Decimal(str(targetSheet['J7'].value))) if targetSheet['J7'].value != 0 else 0
    targetSheet['Q7'] = str(Q7 * Decimal(100))+'%'

    Q12 = ((Decimal(str(targetSheet['L6'].value)) + Decimal(str(targetSheet['L7'].value)))/ (Decimal(str(targetSheet['J6'].value)) + Decimal(str(targetSheet['J7'].value)))) if targetSheet['J6'].value !=0 and targetSheet['J7'].value !=0 else 0 
    targetSheet['Q12'] = str(Q12 * Decimal(100))+'%'

    Q13 = (Decimal(str(targetSheet['L13'].value)) / Decimal(str(targetSheet['J13'].value))) if targetSheet['J13'].value != 0 else 0
    targetSheet['Q13'] = str(Q13 * Decimal(100))+'%'

    Q14 = (Decimal(str(targetSheet['L14'].value)) / Decimal(str(targetSheet['J14'].value))) if targetSheet['J14'].value !=0 else 0
    targetSheet['Q14'] = str(Q14 * Decimal(100))+'%'

    Q19 = ((Decimal(str(targetSheet['L13'].value)) + Decimal(str(targetSheet['L14'].value)))/ (Decimal(str(targetSheet['J13'].value)) + Decimal(str(targetSheet['J14'].value)))) if  targetSheet['J13'].value != 0 and targetSheet['J14'].value!=0 else 0
    targetSheet['Q19'] = str(Q19 * Decimal(100))+'%'

    Q20 = (Decimal(str(targetSheet['L20'].value)) / Decimal(str(targetSheet['J20'].value))) if targetSheet['J20'].value !=0 else 0
    targetSheet['Q20'] = str(Q20 * Decimal(100))+'%'
    
    Q21 = (Decimal(str(targetSheet['L21'].value)) / Decimal(str(targetSheet['J21'].value))) if targetSheet['J21'].value !=0 else 0
    targetSheet['Q21'] = str(Q21 * Decimal(100))+'%'

    Q26 = ((Decimal(str(targetSheet['L20'].value)) + Decimal(str(targetSheet['L21'].value)))/ (Decimal(str(targetSheet['J20'].value)) + Decimal(str(targetSheet['J21'].value)))) if targetSheet['J20'].value != 0 and targetSheet['J21'].value != 0 else 0
    targetSheet['Q26'] = str(Q26 * Decimal(100))+'%'

    template.active.title=codeDict[str(gongsi)]

    template.save('test.xlsx')

    # for k,v in d6.items():
    #     print(k,v)
    print('结束')
    return 1

# entry(wb,'10月')


def outerEntry(targetWb, destnation, keys, dict, name):
    worksheet = targetWb[str(name)]  # 辅助余额表数据源
    entry(targetWb, destnation, name, 1900)  # 每个公司对应的数据
    wb = Workbook()
    # key 是公司代码
    for key in keys:
        pass
        # ws = wb.create_sheet(dict[key])
        # template = entry(targetWb, destnation, name, key, dict[key])  # 每个公司对应的数据
        # print('template==>',template)
        # template.active.title = dict[key]
        # template.save('test-'+dict[key]+'.xlsx')

        # 把对应数据拷贝到 ws 中
        # for row in targetSheet:
        #     for cell in row:
        #         ws[cell.coordinate].value = cell.value

        # wb.save('test.xlsx')

    print('结束')
    return 1