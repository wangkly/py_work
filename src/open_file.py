from tkinter import *
import tkinter.filedialog
from tkinter import ttk
import tkinter.messagebox
from openpyxl import Workbook
from openpyxl import load_workbook
from main import entry
from main import outerEntry
import os

root = Tk()
root.title('徐二狗专用')
root.geometry('500x300')

targetWb = None
destnation = None
submitting = False
gongsi = None

def startwork(name):
    global targetWb #辅助余额表
    global submitting
    global gongsi
    print('submitting',submitting)
    if submitting:
        print('正在执行')
        return
    submitting = True
    if gongsi == None:
        gongsi='1900'
    print('gongsi==>',gongsi)
    result  = entry(targetWb,destnation,name,gongsi)
    if result == 1:
        tkinter.messagebox.showinfo(title='处理成功',message='处理成功')
        # labe2 = Label(root,text="处理成功")
        # labe2.pack()
        submitting = False
        os._exit(0)



def startUseDefaultWork(name):
    global targetWb #辅助余额表
    global submitting
    destWb = load_workbook(str(destnation))
    if submitting:
        print('正在执行')
        return
    submitting = True
    sheet = destWb["filter"]
    dict={}
    keys=[]
    for row in sheet.rows:
        K,V = row
        keys.append(K.value)
        dict[K.value] = V.value
        # xx='{\''+K.value+'\''+":'"+V.value+'\'}'
        # dict.update(eval(xx))
    result = outerEntry(targetWb,destnation,keys,dict,name)
    if result == 1:
        tkinter.messagebox.showinfo(title='处理成功',message='处理成功')
        # labe2 = Label(root,text="处理成功")
        # labe2.pack()
        submitting = False
        os._exit(0)


def placeBtns(item):
    button = Button(root,text=item,command=lambda : startwork(item))
    button.pack()
    return

def combo(*args):
    global gongsi
    print(comboxlist.get())
    code = comboxlist.get()
    if code == '江苏':
        gongsi = '-1'
    else:
        gongsi = code    
    pass


def cmd():
    global targetWb
    filename = tkinter.filedialog.askopenfilename()
    targetWb = load_workbook(filename)
    label2 = Label(root,text='文件路径：'+str(filename))
    label2.pack()
    sheets = targetWb.sheetnames
    label = Label(root,text="第三步:请选择要处理的表格")
    label.pack()
    for item in sheets:
        placeBtns(item)

def cmd2():
    global destnation
    destnation = tkinter.filedialog.askopenfilename()
    label3 = Label(root,text='模板文件路径：'+str(destnation))
    label3.pack()
    

label1=Label(root,text="选择机构")
label1.pack()
 

#加一个下拉框选择要处理的公司
comvalue = tkinter.StringVar()
comboxlist = ttk.Combobox(root,textvariable=comvalue)
comboxlist["values"]=("1900","1902","1903","1904","1905","1906","1999","江苏")
comboxlist.current(0)
comboxlist.bind("<<ComboboxSelected>>",combo)
comboxlist.pack()

btn2 = Button(root,text="第一步：选择要生成文件的模板",command=cmd2)
btn2.pack()

btn = Button(root,text="第二步：选择辅助余额表",command=cmd)
btn.pack()

root.mainloop()
